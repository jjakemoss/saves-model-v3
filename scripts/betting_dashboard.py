"""
Display betting performance dashboard

This script:
1. Reads betting_tracker.xlsx
2. Calculates performance metrics
3. Displays formatted report to console
4. Optionally saves report to file

Usage:
    python scripts/betting_dashboard.py [--save]
"""
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
import argparse

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))

from betting import BettingTracker, calculate_performance_metrics, format_metrics_report


def update_summary_sheet(tracker_file, metrics):
    """
    Update Summary sheet in Excel with calculated metrics

    Args:
        tracker_file: Path to betting tracker Excel file
        metrics: dict from calculate_performance_metrics()
    """
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.load_workbook(tracker_file)
    summary = wb['Summary']

    # Update timestamp
    summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Update Overall Performance section (rows 7-13)
    overall = metrics['overall']
    summary['B7'] = overall['total_bets']
    summary['B8'] = overall['wins']
    summary['B9'] = overall['losses']
    summary['B10'] = overall['pushes']
    summary['B11'] = f"{overall['win_rate']:.1%}"
    summary['B12'] = f"{overall['total_profit_loss']:+.2f}"
    summary['B13'] = f"{overall['roi']:+.2f}%"

    # Update Performance by Confidence section (rows 18-23)
    confidence_buckets = ['50-55%', '55-60%', '60-65%', '65-70%', '70-75%', '75%+']
    for idx, bucket in enumerate(confidence_buckets, 18):
        conf_metrics = metrics['by_confidence'][bucket]
        summary[f'B{idx}'] = conf_metrics['total_bets']
        summary[f'C{idx}'] = conf_metrics['wins']
        summary[f'D{idx}'] = f"{conf_metrics['win_rate']:.1%}" if conf_metrics['total_bets'] > 0 else "0.0%"
        summary[f'E{idx}'] = f"{conf_metrics['roi']:+.2f}%" if conf_metrics['total_bets'] > 0 else "0.0%"

    # Save workbook
    wb.save(tracker_file)
    wb.close()


def show_dashboard(tracker_file='betting_tracker.xlsx', save_report=False):
    """
    Display betting performance dashboard

    Args:
        tracker_file: Path to betting tracker Excel file
        save_report: If True, save report to file

    Returns:
        None
    """
    # Initialize tracker
    tracker = BettingTracker(tracker_file)

    print("\nLoading betting data...")

    # Read all date sheets and combine
    import openpyxl
    wb = openpyxl.load_workbook(tracker_file, read_only=True)
    date_sheets = [s for s in wb.sheetnames if s not in ['Summary', 'Settings']]
    wb.close()

    if len(date_sheets) == 0:
        print("[WARNING] No date sheets found - no betting data available yet")
        return

    # Combine all date sheets
    all_data = []
    for sheet_name in date_sheets:
        df = pd.read_excel(tracker_file, sheet_name=sheet_name)
        all_data.append(df)

    bets_df = pd.concat(all_data, ignore_index=True)

    # Calculate metrics
    print("Calculating performance metrics...\n")
    metrics = calculate_performance_metrics(bets_df)

    # Format report
    report = format_metrics_report(metrics)

    # Display to console
    print(report)

    # Update Summary sheet in Excel
    print("\nUpdating Summary sheet in Excel...")
    update_summary_sheet(tracker_file, metrics)
    print("[OK] Summary sheet updated")

    # Save to file if requested
    if save_report:
        reports_dir = Path('reports')
        reports_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_file = reports_dir / f'betting_performance_{timestamp}.txt'

        with open(report_file, 'w') as f:
            f.write(report)

        print(f"\n[OK] Report saved to: {report_file}")

    # Show insights
    print_insights(metrics)


def print_insights(metrics):
    """
    Print actionable insights from metrics

    Args:
        metrics: dict from calculate_performance_metrics()
    """
    overall = metrics['overall']

    if overall['total_bets'] == 0:
        return

    print("\nKEY INSIGHTS")
    print("-" * 60)

    # Overall performance
    if overall['roi'] > 5:
        print("[OK] Strong overall ROI - strategy is profitable")
    elif overall['roi'] > 0:
        print("[OK] Positive ROI - strategy is working")
    elif overall['roi'] > -5:
        print("[WARNING] Slightly negative ROI - monitor performance")
    else:
        print("[WARNING] Negative ROI - consider strategy adjustment")

    # Sample size
    if overall['total_bets'] < 30:
        print("[WARNING] Small sample size - need more data for reliable conclusions")
    elif overall['total_bets'] < 100:
        print("[INFO] Moderate sample size - trends becoming more reliable")
    else:
        print("[OK] Large sample size - performance metrics are reliable")

    # Confidence bucket analysis
    print("\nConfidence Bucket Performance:")
    best_roi = -999
    best_bucket = None

    for bucket, conf_metrics in metrics['by_confidence'].items():
        if conf_metrics['total_bets'] >= 10:  # Only analyze buckets with enough bets
            roi = conf_metrics['roi']
            if roi > best_roi:
                best_roi = roi
                best_bucket = bucket

            win_rate = conf_metrics['win_rate']
            if roi > 10:
                print(f"  [OK] {bucket}: Excellent performance ({roi:+.1f}% ROI, {win_rate:.1%} win rate)")
            elif roi > 5:
                print(f"  [OK] {bucket}: Strong performance ({roi:+.1f}% ROI, {win_rate:.1%} win rate)")
            elif roi > 0:
                print(f"  [INFO] {bucket}: Profitable ({roi:+.1f}% ROI, {win_rate:.1%} win rate)")
            else:
                print(f"  [WARNING] {bucket}: Unprofitable ({roi:+.1f}% ROI, {win_rate:.1%} win rate)")

    if best_bucket and best_roi > 5:
        print(f"\n[INFO] Best bucket: {best_bucket} - focus bets here")

    # Recent trend
    recent = metrics['recent']
    if 'last_20' in recent and recent['last_20']['total_bets'] >= 10:
        recent_roi = recent['last_20']['roi']
        if recent_roi > overall['roi'] + 5:
            print("\n[OK] Recent performance trending UP")
        elif recent_roi < overall['roi'] - 5:
            print("\n[WARNING] Recent performance trending DOWN")

    # Bet type analysis
    over_metrics = metrics['by_bet_type'].get('OVER', {})
    under_metrics = metrics['by_bet_type'].get('UNDER', {})

    if over_metrics.get('total_bets', 0) >= 10 and under_metrics.get('total_bets', 0) >= 10:
        over_roi = over_metrics['roi']
        under_roi = under_metrics['roi']

        print(f"\nBet Type Preference:")
        if abs(over_roi - under_roi) > 10:
            if over_roi > under_roi:
                print(f"  [INFO] OVER bets performing significantly better ({over_roi:+.1f}% vs {under_roi:+.1f}%)")
            else:
                print(f"  [INFO] UNDER bets performing significantly better ({under_roi:+.1f}% vs {over_roi:+.1f}%)")

    print("")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Display betting performance dashboard'
    )
    parser.add_argument(
        '--tracker',
        type=str,
        default='betting_tracker.xlsx',
        help='Path to betting tracker file. Default: betting_tracker.xlsx'
    )
    parser.add_argument(
        '--save',
        action='store_true',
        help='Save report to reports/ directory'
    )

    args = parser.parse_args()

    try:
        show_dashboard(tracker_file=args.tracker, save_report=args.save)
    except FileNotFoundError as e:
        print(f"\n[ERROR] {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
