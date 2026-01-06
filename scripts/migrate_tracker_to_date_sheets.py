"""
Migrate existing betting_tracker.xlsx from single 'Bets' sheet to date-based sheets

This script:
1. Backs up the existing tracker
2. Reads all data from the 'Bets' sheet
3. Creates new tracker with Summary and Settings sheets
4. Creates date-based sheets for each unique date
5. Migrates all data to the appropriate date sheets

Usage:
    python scripts/migrate_tracker_to_date_sheets.py
"""
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from datetime import datetime
import shutil

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))

from betting import BettingTracker


def migrate_tracker():
    """Migrate existing tracker to date-based sheets format"""

    tracker_file = Path('betting_tracker.xlsx')

    if not tracker_file.exists():
        print("[ERROR] betting_tracker.xlsx not found")
        print("Nothing to migrate")
        return

    # Check if already migrated
    wb = openpyxl.load_workbook(tracker_file, read_only=True)
    has_bets_sheet = 'Bets' in wb.sheetnames
    wb.close()

    if not has_bets_sheet:
        print("[OK] Tracker already uses date-based sheets")
        print("No migration needed")
        return

    print("Migrating betting_tracker.xlsx to date-based sheets format...")

    # Create backup
    backup_file = f'betting_tracker_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    shutil.copy(tracker_file, backup_file)
    print(f"[OK] Backup created: {backup_file}")

    # Read existing data from Bets sheet
    print("Reading existing data from Bets sheet...")
    bets_df = pd.read_excel(tracker_file, sheet_name='Bets')

    # Also read Summary and Settings if they exist
    wb = openpyxl.load_workbook(tracker_file)
    has_summary = 'Summary' in wb.sheetnames
    has_settings = 'Settings' in wb.sheetnames
    wb.close()

    # Create new tracker structure
    print("Creating new tracker structure...")

    # Delete old tracker
    tracker_file.unlink()

    # Run init script to create new structure
    import subprocess
    subprocess.run([sys.executable, 'scripts/init_betting_tracker.py'], check=True)

    # Migrate data if there was any
    if len(bets_df) > 0:
        print(f"\nMigrating {len(bets_df)} games to date-based sheets...")

        # Group by date
        bets_df['game_date'] = pd.to_datetime(bets_df['game_date']).dt.strftime('%Y-%m-%d')
        dates = bets_df['game_date'].unique()

        tracker = BettingTracker(tracker_file)

        for date in sorted(dates):
            date_games = bets_df[bets_df['game_date'] == date]
            print(f"  Creating sheet for {date} ({len(date_games)} games)")

            # Get or create sheet for this date
            sheet_name = tracker._get_or_create_date_sheet(date)

            # Load workbook and write data
            wb = openpyxl.load_workbook(tracker_file)
            ws = wb[sheet_name]

            # Start from row 2 (after header)
            row_num = 2

            for _, game in date_games.iterrows():
                # Write each column
                ws.cell(row=row_num, column=1, value=game['game_date'])
                ws.cell(row=row_num, column=2, value=game['game_id'])
                ws.cell(row=row_num, column=3, value=game['goalie_name'])
                ws.cell(row=row_num, column=4, value=game.get('betting_line', ''))
                ws.cell(row=row_num, column=5, value=game.get('goalie_id', ''))
                ws.cell(row=row_num, column=6, value=game['team_abbrev'])
                ws.cell(row=row_num, column=7, value=game['opponent_team'])
                ws.cell(row=row_num, column=8, value=game['is_home'])
                ws.cell(row=row_num, column=9, value=game.get('predicted_saves', ''))
                ws.cell(row=row_num, column=10, value=game.get('prob_over', ''))
                ws.cell(row=row_num, column=11, value=game.get('confidence_pct', ''))
                ws.cell(row=row_num, column=12, value=game.get('confidence_bucket', ''))
                ws.cell(row=row_num, column=13, value=game.get('recommendation', ''))
                ws.cell(row=row_num, column=14, value=game.get('bet_amount', ''))
                ws.cell(row=row_num, column=15, value=game.get('bet_selection', 'NONE'))
                ws.cell(row=row_num, column=16, value=game.get('actual_saves', ''))
                ws.cell(row=row_num, column=17, value=game.get('result', ''))
                ws.cell(row=row_num, column=18, value=game.get('profit_loss', ''))
                ws.cell(row=row_num, column=19, value=game.get('notes', ''))

                row_num += 1

            wb.save(tracker_file)
            wb.close()

        print(f"\n[OK] Migrated {len(bets_df)} games to {len(dates)} date sheets")

    print("\n" + "="*60)
    print("[OK] Migration complete!")
    print("="*60)
    print(f"\nBackup saved as: {backup_file}")
    print("\nNew structure:")
    print("  1. Summary - Performance metrics (first tab)")
    print("  2. Settings - Configuration (second tab)")
    print("  3. Date sheets - One per day (newest first)")
    print("\nAll scripts will now work with the new date-based format")


if __name__ == '__main__':
    migrate_tracker()
