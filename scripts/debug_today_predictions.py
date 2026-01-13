"""
Debug script to check what predictions were generated today.
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

import pandas as pd
from openpyxl import load_workbook

def check_today_predictions():
    """Check predictions in betting tracker for today."""
    print("="*80)
    print("CHECKING TODAY'S PREDICTIONS")
    print("="*80)
    print()

    # Load the Excel file
    wb = load_workbook('betting_tracker.xlsx')

    # Check if today's sheet exists
    today = '2026-01-12'

    if today not in wb.sheetnames:
        print(f"No sheet found for {today}")
        return

    # Read the sheet
    ws = wb[today]

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]

    # Get all data rows
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):  # Skip empty rows
            data.append(row)

    # Create DataFrame
    df = pd.DataFrame(data, columns=headers)

    print(f"Sheet: {today}")
    print(f"Total rows: {len(df)}")
    print()

    # Check predictions
    if 'prob_over' not in df.columns:
        print("No prob_over column found")
        return

    # Filter to rows with predictions
    predictions = df[df['prob_over'].notna()].copy()

    print(f"Rows with predictions: {len(predictions)}")
    print()

    if len(predictions) == 0:
        print("No predictions found")
        return

    # Analyze predictions
    print("PREDICTION ANALYSIS:")
    print("-"*80)

    for idx, row in predictions.iterrows():
        goalie = row.get('goalie_name', 'Unknown')
        prob_over = row.get('prob_over')
        recommendation = row.get('recommendation', 'N/A')
        ev = row.get('recommended_ev')

        print(f"{goalie}:")
        print(f"  prob_over: {prob_over:.4f} ({prob_over*100:.1f}%)")
        print(f"  recommendation: {recommendation}")
        if ev is not None:
            print(f"  recommended_ev: {ev:.4f} ({ev*100:.1f}%)")
        print()

    # Summary statistics
    prob_over_values = predictions['prob_over'].dropna()

    print("="*80)
    print("SUMMARY STATISTICS:")
    print("-"*80)
    print(f"Total predictions: {len(prob_over_values)}")
    print(f"Predictions >50%: {sum(prob_over_values > 0.5)} ({100*sum(prob_over_values > 0.5)/len(prob_over_values):.1f}%)")
    print(f"Predictions <=50%: {sum(prob_over_values <= 0.5)} ({100*sum(prob_over_values <= 0.5)/len(prob_over_values):.1f}%)")
    print()
    print(f"Mean prob_over: {prob_over_values.mean():.4f}")
    print(f"Median prob_over: {prob_over_values.median():.4f}")
    print(f"Min prob_over: {prob_over_values.min():.4f}")
    print(f"Max prob_over: {prob_over_values.max():.4f}")
    print()

    # Check recommendations
    if 'recommendation' in predictions.columns:
        rec_counts = predictions['recommendation'].value_counts()
        print("Recommendations breakdown:")
        for rec, count in rec_counts.items():
            print(f"  {rec}: {count}")

if __name__ == '__main__':
    check_today_predictions()
