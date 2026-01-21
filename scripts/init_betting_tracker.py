"""
Initialize betting tracker Excel file with proper structure

Creates betting_tracker.xlsx with:
- Summary sheet (performance metrics) - First tab
- Settings sheet (configuration) - Second tab
- Date-based sheets for each day's games (newest first)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime

def create_date_sheet(wb, sheet_name, position):
    """
    Create a date-based sheet with proper headers and formatting

    Args:
        wb: Workbook object
        sheet_name: Name for the sheet (e.g., '2026-01-05')
        position: Position index for sheet

    Returns:
        Worksheet object
    """
    sheet = wb.create_sheet(sheet_name, position)

    # Define columns (book column at position 3)
    headers = [
        'game_date', 'game_id', 'book', 'goalie_name', 'betting_line', 'line_over', 'line_under',
        'goalie_id', 'team_abbrev', 'opponent_team', 'is_home', 'predicted_saves',
        'prob_over', 'confidence_pct', 'confidence_bucket', 'recommendation', 'ev',
        'bet_amount', 'bet_selection', 'actual_saves', 'result',
        'profit_loss', 'notes'
    ]

    # Header styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths (with book column at C)
    column_widths = {
        'A': 12,  # game_date
        'B': 12,  # game_id
        'C': 12,  # book
        'D': 18,  # goalie_name
        'E': 13,  # betting_line
        'F': 12,  # line_over
        'G': 12,  # line_under
        'H': 12,  # goalie_id
        'I': 12,  # team_abbrev
        'J': 15,  # opponent_team
        'K': 10,  # is_home
        'L': 15,  # predicted_saves
        'M': 12,  # prob_over
        'N': 14,  # confidence_pct
        'O': 16,  # confidence_bucket
        'P': 15,  # recommendation
        'Q': 12,  # ev
        'R': 12,  # bet_amount
        'S': 14,  # bet_selection
        'T': 13,  # actual_saves
        'U': 10,  # result
        'V': 12,  # profit_loss
        'W': 30,  # notes
    }

    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Freeze top row
    sheet.freeze_panes = 'A2'

    return sheet

def create_betting_tracker():
    """Create new betting tracker Excel file"""

    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create Summary sheet (first tab)
    summary_sheet = wb.create_sheet('Summary', 0)

    # Summary headers
    summary_sheet['A1'] = 'BETTING PERFORMANCE SUMMARY'
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet.merge_cells('A1:D1')

    summary_sheet['A3'] = 'Generated:'
    summary_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Overall Performance section
    summary_sheet['A5'] = 'OVERALL PERFORMANCE'
    summary_sheet['A5'].font = Font(bold=True, size=12)

    # Note: Formulas will aggregate across all date sheets
    summary_sheet['A7'] = 'Total Bets'
    summary_sheet['A8'] = 'Wins'
    summary_sheet['A9'] = 'Losses'
    summary_sheet['A10'] = 'Pushes'
    summary_sheet['A11'] = 'Win Rate'
    summary_sheet['A12'] = 'Total Profit/Loss'
    summary_sheet['A13'] = 'ROI %'

    # Placeholder values - will be updated by dashboard script
    for idx in range(7, 14):
        summary_sheet[f'B{idx}'] = 0

    # Performance by Confidence section
    summary_sheet['A16'] = 'PERFORMANCE BY CONFIDENCE LEVEL'
    summary_sheet['A16'].font = Font(bold=True, size=12)

    conf_headers = ['Confidence', 'Bets', 'Wins', 'Win Rate', 'ROI %']
    for col_idx, header in enumerate(conf_headers, 1):
        cell = summary_sheet.cell(row=17, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)

    confidence_levels = ['50-55%', '55-60%', '60-65%', '65-70%', '70-75%', '75%+']
    for idx, conf_level in enumerate(confidence_levels, 18):
        summary_sheet[f'A{idx}'] = conf_level
        # Placeholder values - will be updated by dashboard script
        for col in ['B', 'C', 'D', 'E']:
            summary_sheet[f'{col}{idx}'] = 0

    # Create Settings sheet (second tab)
    settings_sheet = wb.create_sheet('Settings', 1)

    settings_sheet['A1'] = 'BETTING TRACKER SETTINGS'
    settings_sheet['A1'].font = Font(bold=True, size=14)

    settings = [
        ('Default Unit Size', 1.0),
        ('Min Confidence to Bet (%)', 55),
        ('High Confidence Threshold (%)', 65),
        ('Min Expected Value (%)', 2.0),
        ('', ''),
        ('Odds Format', 'American'),
        ('Default Odds (if missing)', -110),
        ('', ''),
        ('Model Path', 'models/classifier_model.json'),
        ('Feature Count', 89),
    ]

    settings_sheet['A3'] = 'Setting'
    settings_sheet['B3'] = 'Value'
    settings_sheet['A3'].font = Font(bold=True)
    settings_sheet['B3'].font = Font(bold=True)

    for idx, (setting, value) in enumerate(settings, 4):
        settings_sheet[f'A{idx}'] = setting
        settings_sheet[f'B{idx}'] = value

    settings_sheet.column_dimensions['A'].width = 30
    settings_sheet.column_dimensions['B'].width = 20

    # Save workbook
    output_path = Path('betting_tracker.xlsx')
    wb.save(output_path)

    print(f'[OK] Created betting tracker: {output_path.absolute()}')
    print('\nSheets created:')
    print('  1. Summary - Performance metrics (first tab)')
    print('  2. Settings - Configuration (second tab)')
    print('  3. Date sheets will be created automatically when games are populated')
    print('\nNext steps:')
    print('  1. Run: python scripts/populate_daily_games.py')
    print('  2. Enter betting lines in Excel')
    print('  3. Run: python scripts/generate_predictions.py')

if __name__ == '__main__':
    create_betting_tracker()
