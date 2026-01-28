"""
Excel manager for betting tracker operations
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from pathlib import Path
from datetime import datetime
import shutil


class BettingTracker:
    """Manage betting tracker Excel file"""

    def __init__(self, tracker_file='betting_tracker.xlsx'):
        """
        Initialize tracker manager

        Args:
            tracker_file: Path to betting tracker Excel file
        """
        self.tracker_file = Path(tracker_file)

        if not self.tracker_file.exists():
            raise FileNotFoundError(
                f"Betting tracker not found: {tracker_file}\n"
                f"Run: python scripts/init_betting_tracker.py"
            )

    def _get_or_create_date_sheet(self, date_str):
        """
        Get existing date sheet or create new one

        Args:
            date_str: Date string (YYYY-MM-DD)

        Returns:
            str: Sheet name
        """
        wb = openpyxl.load_workbook(self.tracker_file)

        # Check if sheet exists
        if date_str in wb.sheetnames:
            wb.close()
            return date_str

        # Create new sheet - insert after Settings (position 2)
        # Find the right position (after all newer dates)
        existing_dates = [s for s in wb.sheetnames if s not in ['Summary', 'Settings']]
        existing_dates.sort(reverse=True)  # Newest first

        # Find position to insert
        position = 2  # After Summary and Settings
        for existing_date in existing_dates:
            if existing_date > date_str:
                position += 1
            else:
                break

        sheet = wb.create_sheet(date_str, position)

        # Define columns (book column added at position 3)
        headers = [
            'game_date', 'game_id', 'book', 'goalie_name', 'betting_line', 'line_over', 'line_under',
            'goalie_id', 'team_abbrev', 'opponent_team', 'is_home', 'predicted_saves',
            'prob_over', 'confidence_pct', 'confidence_bucket', 'recommendation', 'ev',
            'bet_amount', 'bet_selection', 'actual_saves', 'result',
            'profit_loss', 'notes'
        ]

        # Header styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Set column widths (with book column at C)
        column_widths = {
            'A': 12,  # game_date
            'B': 12,  # game_id
            'C': 12,  # book
            'D': 18,  # goalie_name
            'E': 13,  # betting_line
            'F': 12,  # line_over
            'G': 12,  # line_under
            'H': 12,  # goalie_id
            'I': 12,  # team_abbrev
            'J': 15,  # opponent_team
            'K': 10,  # is_home
            'L': 15,  # predicted_saves
            'M': 12,  # prob_over
            'N': 14,  # confidence_pct
            'O': 16,  # confidence_bucket
            'P': 15,  # recommendation
            'Q': 12,  # ev
            'R': 12,  # bet_amount
            'S': 14,  # bet_selection
            'T': 13,  # actual_saves
            'U': 10,  # result
            'V': 12,  # profit_loss
            'W': 30,  # notes
        }

        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        # Freeze top row
        sheet.freeze_panes = 'A2'

        # Save and close
        wb.save(self.tracker_file)
        wb.close()

        return date_str

    def append_games(self, games_df):
        """
        Append new game rows to appropriate date sheet

        Args:
            games_df: pd.DataFrame with game info (partial columns)
        """
        if len(games_df) == 0:
            return

        # Group games by date
        games_by_date = games_df.groupby('game_date')

        total_appended = 0

        for game_date, date_games in games_by_date:
            # Get or create sheet for this date
            sheet_name = self._get_or_create_date_sheet(game_date)

            # Load workbook
            wb = openpyxl.load_workbook(self.tracker_file)
            ws = wb[sheet_name]

            # Find next empty row
            next_row = ws.max_row + 1

            # Append each game
            for _, game in date_games.iterrows():
                row_data = [
                    game.get('game_date'),
                    game.get('game_id'),
                    game.get('book', ''),  # book (e.g., 'Underdog')
                    game.get('goalie_name'),
                    game.get('betting_line', ''),  # betting_line (pre-filled or user fills)
                    game.get('line_over', ''),  # line_over (pre-filled or user fills)
                    game.get('line_under', ''),  # line_under (pre-filled or user fills)
                    game.get('goalie_id'),
                    game.get('team_abbrev'),
                    game.get('opponent_team'),
                    game.get('is_home'),
                    '',  # predicted_saves
                    '',  # prob_over
                    '',  # confidence_pct
                    '',  # confidence_bucket
                    '',  # recommendation
                    '',  # ev
                    '',  # bet_amount
                    'NONE',  # bet_selection
                    '',  # actual_saves
                    '',  # result
                    '',  # profit_loss
                    '',  # notes
                ]

                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=next_row, column=col_idx, value=value)

                next_row += 1
                total_appended += 1

            # Save
            wb.save(self.tracker_file)
            wb.close()

        print(f"[OK] Appended {total_appended} games to {self.tracker_file}")

    def update_predictions(self, predictions_df):
        """
        Update prediction columns for games

        Args:
            predictions_df: pd.DataFrame with game_id, game_date, and prediction columns
        """
        if len(predictions_df) == 0:
            return

        # Group predictions by date
        preds_by_date = predictions_df.groupby('game_date')

        total_updated = 0

        for game_date, date_preds in preds_by_date:
            # Get sheet for this date
            sheet_name = self._get_or_create_date_sheet(game_date)

            # Load existing data with explicit dtypes
            df = pd.read_excel(self.tracker_file, sheet_name=sheet_name, dtype={
                'result': 'object',
                'confidence_bucket': 'object',
                'recommendation': 'object',
                'bet_selection': 'object',
            })

            # Update predictions
            for _, pred in date_preds.iterrows():
                game_id = pred['game_id']
                goalie_id = pred['goalie_id']
                goalie_name = pred.get('goalie_name', '')
                book = pred.get('book', '')
                betting_line = pred.get('betting_line')
                line_over = pred.get('line_over')
                line_under = pred.get('line_under')

                # Match by game_id + goalie_name + book + betting_line + odds for precise targeting
                mask = (df['game_id'] == game_id) & (df['goalie_name'].fillna('').astype(str).str.lower() == goalie_name.lower())
                if book and 'book' in df.columns:
                    mask = mask & (df['book'] == book)
                if betting_line is not None and 'betting_line' in df.columns:
                    mask = mask & (df['betting_line'] == betting_line)
                if line_over is not None and 'line_over' in df.columns:
                    mask = mask & (df['line_over'] == line_over)
                if line_under is not None and 'line_under' in df.columns:
                    mask = mask & (df['line_under'] == line_under)

                if mask.any():
                    # Update goalie_id if it was looked up
                    if pd.notna(goalie_id):
                        df.loc[mask, 'goalie_id'] = goalie_id

                    df.loc[mask, 'predicted_saves'] = pred.get('predicted_saves')
                    df.loc[mask, 'prob_over'] = pred.get('prob_over')
                    df.loc[mask, 'confidence_pct'] = pred.get('confidence_pct')
                    df.loc[mask, 'confidence_bucket'] = pred.get('confidence_bucket')
                    df.loc[mask, 'recommendation'] = pred.get('recommendation')
                    df.loc[mask, 'ev'] = pred.get('recommended_ev')
                    total_updated += mask.sum()

            # Write back to Excel
            with pd.ExcelWriter(self.tracker_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"[OK] Updated {total_updated} predictions")

    def update_results(self, results_df):
        """
        Update actual saves and results for completed games

        Args:
            results_df: pd.DataFrame with game_id, game_date, actual_saves, result, profit_loss
        """
        if len(results_df) == 0:
            return

        # Group results by date
        results_by_date = results_df.groupby('game_date')

        total_updated = 0

        for game_date, date_results in results_by_date:
            # Get sheet for this date
            sheet_name = self._get_or_create_date_sheet(game_date)

            # Load existing data with explicit dtypes
            df = pd.read_excel(self.tracker_file, sheet_name=sheet_name, dtype={
                'result': 'object',
                'confidence_bucket': 'object',
                'recommendation': 'object',
                'bet_selection': 'object',
            })

            # Update results
            for _, result in date_results.iterrows():
                game_id = result['game_id']
                goalie_id = result['goalie_id']
                book = result.get('book', '')
                betting_line = result.get('betting_line')
                line_over = result.get('line_over')
                line_under = result.get('line_under')

                # Match by game_id + goalie_id + book + betting_line + odds for precise targeting
                mask = (df['game_id'] == game_id) & (df['goalie_id'] == goalie_id)
                if book and 'book' in df.columns:
                    mask = mask & (df['book'] == book)
                if betting_line is not None and 'betting_line' in df.columns:
                    mask = mask & (df['betting_line'] == betting_line)
                if line_over is not None and 'line_over' in df.columns:
                    mask = mask & (df['line_over'] == line_over)
                if line_under is not None and 'line_under' in df.columns:
                    mask = mask & (df['line_under'] == line_under)

                if mask.any():
                    df.loc[mask, 'actual_saves'] = result.get('actual_saves')
                    df.loc[mask, 'result'] = result.get('result')
                    df.loc[mask, 'profit_loss'] = result.get('profit_loss')
                    total_updated += mask.sum()

            # Write back
            with pd.ExcelWriter(self.tracker_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"[OK] Updated results for {total_updated} games")

    def get_todays_games(self, date=None):
        """
        Get games for a specific date from tracker

        Args:
            date: Date string (YYYY-MM-DD). If None, uses today

        Returns:
            pd.DataFrame: Games for that date
        """
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')

        # Check if sheet exists for this date
        wb = openpyxl.load_workbook(self.tracker_file, read_only=True)
        if date not in wb.sheetnames:
            wb.close()
            return pd.DataFrame()

        wb.close()

        # Read the date sheet
        df = pd.read_excel(self.tracker_file, sheet_name=date)
        return df

    def get_pending_predictions(self, date=None):
        """
        Get games that need predictions (have goalie name and betting line but no prediction)

        Args:
            date: Date string. If None, uses today

        Returns:
            pd.DataFrame: Games needing predictions
        """
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')

        # Check if sheet exists for this date
        wb = openpyxl.load_workbook(self.tracker_file, read_only=True)
        if date not in wb.sheetnames:
            wb.close()
            return pd.DataFrame()

        wb.close()

        # Read the date sheet
        df = pd.read_excel(self.tracker_file, sheet_name=date)

        # Filter: has goalie name, has betting line, no prediction yet
        mask = (
            (df['goalie_name'].notna()) &
            (df['goalie_name'] != '') &
            (df['goalie_name'] != 'TBD') &
            (df['betting_line'].notna()) &
            (df['betting_line'] != '') &
            ((df['predicted_saves'].isna()) | (df['predicted_saves'] == ''))
        )

        return df[mask]

    def get_pending_results(self, date=None):
        """
        Get games needing results update

        Args:
            date: Date string. If None, uses yesterday

        Returns:
            pd.DataFrame: Games needing results
        """
        if date is None:
            yesterday = datetime.now() - pd.Timedelta(days=1)
            date = yesterday.strftime('%Y-%m-%d')

        # Check if sheet exists for this date
        wb = openpyxl.load_workbook(self.tracker_file, read_only=True)
        if date not in wb.sheetnames:
            wb.close()
            return pd.DataFrame()

        wb.close()

        # Read the date sheet
        df = pd.read_excel(self.tracker_file, sheet_name=date)

        # Filter: no results yet
        mask = ((df['actual_saves'].isna()) | (df['actual_saves'] == ''))

        return df[mask]

    def backup_to_csv(self, backup_dir='data/betting_history'):
        """
        Save backup of all date sheets to CSV

        Args:
            backup_dir: Directory for backups
        """
        backup_path = Path(backup_dir)
        backup_path.mkdir(parents=True, exist_ok=True)

        # Get all date sheets
        wb = openpyxl.load_workbook(self.tracker_file, read_only=True)
        date_sheets = [s for s in wb.sheetnames if s not in ['Summary', 'Settings']]
        wb.close()

        # Combine all date sheets
        all_data = []
        for sheet_name in date_sheets:
            df = pd.read_excel(self.tracker_file, sheet_name=sheet_name)
            all_data.append(df)

        if all_data:
            combined_df = pd.concat(all_data, ignore_index=True)

            # Save with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            csv_file = backup_path / f'betting_tracker_{timestamp}.csv'
            combined_df.to_csv(csv_file, index=False)

            print(f"[OK] Backup saved: {csv_file}")
            return csv_file

        return None
