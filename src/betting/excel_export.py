"""
Generate the read-only betting_tracker.xlsx snapshot from data/betting.db.

This file is regenerated after every write to the database (new lines,
predictions, results, or a recorded bet). It exists purely for browsing --
it should never be hand-edited. All writes go through BettingDB instead.
"""
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime

from .db_manager import DEFAULT_DB_PATH
from .metrics import calculate_performance_metrics

DEFAULT_XLSX_PATH = 'betting_tracker.xlsx'

BET_COLUMNS = [
    'game_date', 'game_id', 'book', 'goalie_name', 'betting_line', 'line_over', 'line_under',
    'goalie_id', 'team_abbrev', 'opponent_team', 'is_home', 'predicted_saves',
    'prob_over', 'confidence_pct', 'confidence_bucket', 'recommendation', 'ev',
    'bet_amount', 'bet_selection', 'bet_placed_at', 'actual_saves', 'result',
    'profit_loss', 'notes'
]

COLUMN_WIDTHS = {
    'A': 12, 'B': 12, 'C': 12, 'D': 18, 'E': 13, 'F': 12, 'G': 12,
    'H': 12, 'I': 12, 'J': 15, 'K': 10, 'L': 15, 'M': 12, 'N': 14,
    'O': 16, 'P': 15, 'Q': 12, 'R': 12, 'S': 14, 'T': 18, 'U': 13,
    'V': 10, 'W': 12, 'X': 30,
}

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

CONFIDENCE_BUCKETS = ['50-55%', '55-60%', '60-65%', '65-70%', '70-75%', '75%+']


def _style_header_row(sheet, num_columns):
    for col_idx in range(1, num_columns + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
    for col, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = 'A2'


def _write_date_sheet(wb, date_str, position, date_df):
    sheet = wb.create_sheet(date_str, position)

    for col_idx, header in enumerate(BET_COLUMNS, 1):
        sheet.cell(row=1, column=col_idx, value=header)

    for row_idx, (_, row) in enumerate(date_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(BET_COLUMNS, 1):
            value = row.get(col_name)
            if pd.isna(value):
                value = None
            sheet.cell(row=row_idx, column=col_idx, value=value)

    _style_header_row(sheet, len(BET_COLUMNS))
    return sheet


def _write_summary_sheet(wb, metrics):
    sheet = wb.create_sheet('Summary', 0)

    sheet['A1'] = 'BETTING PERFORMANCE SUMMARY'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet.merge_cells('A1:D1')

    sheet['A3'] = 'Generated:'
    sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet['A5'] = 'OVERALL PERFORMANCE'
    sheet['A5'].font = Font(bold=True, size=12)

    overall = metrics['overall']
    labels = [
        ('Total Bets', overall['total_bets']),
        ('Wins', overall['wins']),
        ('Losses', overall['losses']),
        ('Pushes', overall['pushes']),
        ('Win Rate', f"{overall['win_rate']:.1%}"),
        ('Total Profit/Loss', f"{overall['total_profit_loss']:+.2f}"),
        ('ROI %', f"{overall['roi']:+.2f}%"),
    ]
    for idx, (label, value) in enumerate(labels, 7):
        sheet[f'A{idx}'] = label
        sheet[f'B{idx}'] = value

    sheet['A16'] = 'PERFORMANCE BY CONFIDENCE LEVEL'
    sheet['A16'].font = Font(bold=True, size=12)

    conf_headers = ['Confidence', 'Bets', 'Wins', 'Win Rate', 'ROI %']
    for col_idx, header in enumerate(conf_headers, 1):
        cell = sheet.cell(row=17, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)

    for idx, bucket in enumerate(CONFIDENCE_BUCKETS, 18):
        conf_metrics = metrics['by_confidence'][bucket]
        sheet[f'A{idx}'] = bucket
        sheet[f'B{idx}'] = conf_metrics['total_bets']
        sheet[f'C{idx}'] = conf_metrics['wins']
        sheet[f'D{idx}'] = f"{conf_metrics['win_rate']:.1%}" if conf_metrics['total_bets'] > 0 else '0.0%'
        sheet[f'E{idx}'] = f"{conf_metrics['roi']:+.2f}%" if conf_metrics['total_bets'] > 0 else '0.0%'

    sheet.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = 14

    return sheet


def _write_settings_sheet(wb):
    sheet = wb.create_sheet('Settings', 1)

    sheet['A1'] = 'BETTING TRACKER SETTINGS'
    sheet['A1'].font = Font(bold=True, size=14)

    settings = [
        ('Default Unit Size', 1.0),
        ('Minimum EV Threshold (%)', 12.0),
        ('Odds Format', 'American'),
        ('Default Odds (if missing)', -110),
        ('', ''),
        ('Model', 'tuned_v1_20260201_155204'),
        ('Model Path', 'models/trained/tuned_v1_20260201_155204/classifier_model.json'),
        ('Feature Count', 114),
        ('', ''),
        ('Source of Truth', 'data/betting.db'),
        ('This File', 'Read-only snapshot -- regenerated automatically, do not hand-edit'),
    ]

    sheet['A3'] = 'Setting'
    sheet['B3'] = 'Value'
    sheet['A3'].font = Font(bold=True)
    sheet['B3'].font = Font(bold=True)

    for idx, (setting, value) in enumerate(settings, 4):
        sheet[f'A{idx}'] = setting
        sheet[f'B{idx}'] = value

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 55

    return sheet


def export_to_excel(db_path=DEFAULT_DB_PATH, xlsx_path=DEFAULT_XLSX_PATH):
    """
    Regenerate the read-only Excel snapshot from the SQLite database.

    Safe to call any time -- always overwrites xlsx_path with a fresh
    rendering of the current database state.

    Args:
        db_path: Path to the SQLite database
        xlsx_path: Path to write the Excel snapshot to

    Returns:
        Path: The xlsx_path that was written
    """
    db_path = Path(db_path)
    xlsx_path = Path(xlsx_path)

    conn = sqlite3.connect(db_path)
    try:
        bets_df = pd.read_sql_query("SELECT * FROM bets ORDER BY game_date, id", conn)
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    metrics_source = bets_df if len(bets_df) else pd.DataFrame(columns=BET_COLUMNS)
    metrics = calculate_performance_metrics(metrics_source)

    _write_summary_sheet(wb, metrics)
    _write_settings_sheet(wb)

    if len(bets_df):
        dates = sorted(bets_df['game_date'].dropna().unique(), reverse=True)
        for position, date_str in enumerate(dates, start=2):
            date_df = bets_df[bets_df['game_date'] == date_str]
            _write_date_sheet(wb, date_str, position, date_df)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    wb.close()

    return xlsx_path
